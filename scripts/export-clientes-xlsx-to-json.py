"""Export aba CLIENTES do Excel para data/acompanhamento-clientes.json"""
from __future__ import annotations

import json
from pathlib import Path

from openpyxl import load_workbook

XLSX = Path(r"C:\Users\admin\Downloads\CLIENTES 2026 (1) (1).xlsx")
OUT = Path(__file__).resolve().parents[1] / "data" / "acompanhamento-clientes.json"
SHEET_NAME = "CLIENTES"


def cell_str(value) -> str:
    if value is None:
        return ""
    if hasattr(value, "strftime"):
        try:
            return value.strftime("%d/%m/%Y")
        except Exception:
            return str(value).strip()
    return str(value).strip()


def main() -> None:
    wb = load_workbook(XLSX, data_only=True, read_only=True)
    if SHEET_NAME not in wb.sheetnames:
        raise SystemExit(f"Aba {SHEET_NAME!r} não encontrada. Abas: {wb.sheetnames}")

    ws = wb[SHEET_NAME]
    rows_iter = ws.iter_rows(values_only=True)
    header_row = next(rows_iter, None)
    if not header_row:
        raise SystemExit("Planilha vazia")

    headers = [cell_str(h) for h in header_row]
    # trim trailing empty headers
    while headers and not headers[-1]:
        headers.pop()

    out_rows: list[list[str]] = []
    for row in rows_iter:
        cells = [cell_str(row[i]) if i < len(row) else "" for i in range(len(headers))]
        if not any(cells):
            continue
        # precisa ter nome
        if not cells[0]:
            continue
        out_rows.append(cells)

    payload = {"sheet": SHEET_NAME, "headers": headers, "rows": out_rows}
    OUT.write_text(json.dumps(payload, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    print(f"OK: {len(out_rows)} clientes -> {OUT}")


if __name__ == "__main__":
    main()
